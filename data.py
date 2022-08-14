#-------------------------------------------------------------------------------
# Name:        Benes Assignment Program
# Purpose:     Dieses Tool erstellt Vorschläge für die Einteilung von Teachers für verschiedene Kurse.
#
# Author:      Benedikt Schenk
# Copyright:   © 2022 Benedikt Schenk
#-------------------------------------------------------------------------------

import sys
from openpyxl import load_workbook
from openpyxl.styles import Font
import geopy.distance
import openpyxl
from openpyxl.styles import Font



#TODO: make names variable
availabilities_sheet_name = "Übersicht Terminwünsche"
desired_workdays_sheet_name = "Anzahl Wunschtage"
teacher_coordinates_sheet_name = "guides"
event_coordinate_sheet_name = "events"



def get_problem(availability_filepath, coordinates_filepath):

    # import excel sheets
    workbook1 = load_workbook(availability_filepath, data_only=True)
    availabilities_sheet = workbook1[availabilities_sheet_name]
    desired_workdays_sheet = workbook1[desired_workdays_sheet_name]

    workbook2 = load_workbook(coordinates_filepath, data_only=True)
    teacher_coordinates_sheet = workbook2[teacher_coordinates_sheet_name]
    event_coordinates_sheet = workbook2[event_coordinate_sheet_name]

    # dictionary with lat/long coordinates for each teacher
    teacher_coordinates_dict = {row[0][0] : row[0][1] for row in list(zip(teacher_coordinates_sheet.values))[2:]}

    # dictionary with lat/long coordinates for each event
    event_coordinates_dict = {}
    event_coordinate_sheet_as_dict = {row[0][0] : row[0][1] for row in list(zip(event_coordinates_sheet.values))[2:]}
    for row in availabilities_sheet.iter_rows(min_row=3, min_col=1, max_col=2):
        location = row[1].value
        if not location: continue 
        if not location in event_coordinate_sheet_as_dict:
            print("Error: Location \'" + str(location) + "\' could not be found in the Event Location Sheet")
            sys.exit()
        coordinate = event_coordinate_sheet_as_dict[location]
        event_coordinates_dict[row[0].value] = coordinate


    # data initialation
    teachers, events, event_overlap_sets = [],[],[]
    desired_workdays, event_size, event_durations = {},{},{}

    for row in desired_workdays_sheet.iter_rows(min_row=3, min_col=1, max_col=2):
        teacher = row[0].value
        if not teacher: break
        teachers.append(teacher)                         # names of teachers

        if(row[1].value > 0):
            desired_workdays[teacher] = row[1].value     # how many days each teacher wants to work
        else:
            print("Error: Please exclude teachers who don't wish to work.")
            sys.exit()


    for row in availabilities_sheet.iter_rows(min_row=3, min_col=1, max_col=5):
        event = row[0].value
        if not event: continue
        events.append(event)                         # event IDs
        event_durations[event] = row[2].value        # how many days an event lasts
        event_size[event] = row[3].value             # how many teachers are needed per event


        # read overlaps
        overlaps = []
        if type(row[4].value) is str:               
            overlaps = list(map(int, row[4].value.split(";")))
        elif type(row[4].value) is int:
            overlaps = [row[4].value]  
        elif row[4].value is None:
            continue
        else:
            print(f"Error: Could not identify overlapping event \'{row[4].value}\'. Please use \';\' to seperate event ids.")
            sys.exit()

        # find overlapping events and put according event IDs into groups (sets)
        new_group = set()
        for overlap_id in list(filter(lambda x: event < x, overlaps)):
            group_found = False
            for group in event_overlap_sets:
                if overlap_id in group:
                    group.add(event)
                    group_found = True
            if not group_found:
                new_group.add(overlap_id)

        if new_group:
            new_group.add(event)
            event_overlap_sets.append(new_group)

    # an assignment is a tuple (teacher, event, priority, distance), 
    # eg: ("A", 3, 1, 10) means teacher A is assigned to event 3, which he entered as priority 1 and has 10km of distance to
    possible_assignments = []
    for prio in range(4):

        for event, available_teachers in zip(events, availabilities_sheet.iter_rows(min_row=3, min_col=7+prio, max_col=7+prio)):
        
            if available_teachers[0].value:
                for teacher in available_teachers[0].value.replace(" ", "").split(","):
                    if not teacher == "":
                        if not teacher in teachers:
                            print("Error: The teacher-name \'" + teacher + "\' does not seem to match any teacher in the 'desired assignments' sheet")
                            sys.exit()

                        coordinate_teacher = tuple(map(float, teacher_coordinates_dict[teacher].split(",")))
                        coordinate_event = tuple(map(float, event_coordinates_dict[event].split(",")))
                        distance = geopy.distance.distance(coordinate_teacher, coordinate_event).km * 2

                        possible_assignments.append((teacher, event, prio, distance))


    if(len(teachers) > 20):
        print("Error: Maximale Anzahl an teachers überschritten.")
        sys.exit()
    if(len(events) > 40):
        print("Error: Maximale Anzahl an Events überschritten.")

    return teachers, events, event_overlap_sets, desired_workdays, event_size, event_durations, possible_assignments



def get_cost(assignments, worst_distance, weights, teachers, desired_workdays, event_durations, event_size):
            a1, a2, a3, a4 = [w / max(list(map(abs, weights))) / len(weights) for w in weights] # balance weights (highest weight = 1/4 or -1/4)
            num_assigned_workdays = {t : sum([event_durations[a[1]] for a in assignments if a[0] == t]) for t in teachers}
            absolute_workday_deviations = {t : abs(num_assigned_workdays[t] - desired_workdays[t]) for t in teachers}
            rel_workday_deviations = {t : v / desired_workdays[t] for t, v in absolute_workday_deviations.items()}
            average_rel_workday_deviation = sum(rel_workday_deviations.values()) / len(rel_workday_deviations)
            overall_distance = sum([a[3] for a in assignments])
            sum_priorities = sum([a[2] for a in assignments])
            DELTA = max(rel_workday_deviations.values())

            # print('weights:', a1, a2, a3, a4)
            # print('\naverage rel deviation', a1 * average_rel_workday_deviation)
            # print('maximum rel deviation',a2 * DELTA)
            # print('distance', a3 * overall_distance / worst_distance)
            # print('priorities', a4 * (1 - ((sum_priorities / sum(event_size)) / 4)))
            
            costs = [
                a1 * average_rel_workday_deviation + a2 * DELTA,
                a3 * overall_distance / worst_distance,
                a4 * (1 - ((sum_priorities / sum(event_size)) / 4))
            ]

            return costs



def write_model(output_filename, teachers, events, desired_workdays, event_size, event_durations, assignments, weights, worst_case_distance):

    final_assignments_dict_event = {}   # event as key, assignments for this event as value    
    final_assignments_dict_teacher = {}   # same but teacher as key

    for assignment in assignments:
        teacher, event = assignment[0], assignment[1]

        if not teacher in final_assignments_dict_teacher:
            final_assignments_dict_teacher[teacher] = []
        
        if not event in final_assignments_dict_event:
            final_assignments_dict_event[event] = []
        
        final_assignments_dict_event[event].append(assignment)
        final_assignments_dict_teacher[teacher].append(assignment)     


    BLACK = '00000000'
    RED = '00FF0000'
    GREEN = '00008000'
    BLUE = '000000FF'
    ORANGE = 'FFA500'

    solution_sheet_name="Solution Sheet"
    statistics_sheet_name = "Statistics Sheet"

    workbook = openpyxl.Workbook()
    del workbook[workbook.active.title]

    workbook.create_sheet(title=solution_sheet_name)
    solution_sheet = workbook[solution_sheet_name]
    workbook.create_sheet(title=statistics_sheet_name)
    statistics_sheet = workbook[statistics_sheet_name]
    
    c1 = solution_sheet["A1"]
    c2 = solution_sheet["C1"]
    c1.value = "Event"
    c2.value = "teachers"
    c1.font = c2.font = Font(bold=True)

    for curr_row, event in enumerate(events): 

        event_cell = solution_sheet.cell(row=curr_row+3, column=1)
        event_cell.value = event

        num_assigned_teachers = 0

        # write assigned teacher names
        if event in final_assignments_dict_event:
            for teacher_num, assignment in enumerate(final_assignments_dict_event[event]):
                num_assigned_teachers += 1
                teacher_cell = solution_sheet.cell(row=curr_row+3, column=teacher_num+3)
                teacher_cell.value = assignment[0]

                # change color according to priority
                if(assignment[2] == 0):
                    cell_color=ORANGE
                elif(assignment[2] == 1):
                    cell_color=GREEN
                elif(assignment[2] == 2):
                    cell_color=BLUE
                else:
                    cell_color=BLACK

                teacher_cell.font = Font(color=cell_color)
        
        # mark empty spots
        for i in range(num_assigned_teachers, event_size[event]):
            teacher_cell = solution_sheet.cell(row=curr_row+3, column=3+i)
            teacher_cell.value = 'XXX'
            teacher_cell.font = Font(color=RED)



    def get_num_assignments_with_preference(preference):
        return sum(a[2] == preference for a in assignments)


    def write(row, column, value, bold=False):
        c = statistics_sheet.cell(row=row, column=column)
        c.value = value
        c.font=Font(bold=bold)
    
    write(1,1,"teacher", True)
    write(1,2,"Arbeitstage", True)
    write(1,3,"Gewünschte Tage", True)
    write(1,4,"Abweichung in %", True)


    first_teacher_row = 3   
    abs_deviations = []
    relative_deviations = []
    for i, teacher in enumerate(teachers):
        num_workdays = 0
        if(teacher in final_assignments_dict_teacher):
            for assignment in final_assignments_dict_teacher[teacher]:
                num_workdays += event_durations[assignment[1]]

        deviation = num_workdays - desired_workdays[teacher]
        abs_deviations.append(abs(deviation))

        deviation_percentage = deviation / desired_workdays[teacher]
        relative_deviations.append(abs(deviation_percentage))

        write(first_teacher_row+i,1,teacher)
        write(first_teacher_row+i,2,num_workdays)
        write(first_teacher_row+i,3,desired_workdays[teacher])
        write(first_teacher_row+i,4,str(round(100 * deviation_percentage)) + "%")


    overall_distance = sum([assignment[3] for assignment in assignments])
    row = len(teachers) + 5

    write(row, 1, "Durchschnittl Abweichung")
    write(row, 4, str(round(100 * sum(relative_deviations) / len(relative_deviations))) + '%')

    row += 1
    write(row, 1, "Durchschnittl. abs Abweichung")
    write(row, 4, round(sum(abs_deviations) / len(abs_deviations), 1))

    row += 1
    write(row, 1, "Maxim. Abweichung")
    write(row, 4, str(round(100 * max(relative_deviations))) + "%")

    row += 2
    write(row, 1, "Gefahrene km:")
    write(row, 4, round(overall_distance))

    row += 1
    write(row, 1, "Gesparte km:")
    write(row, 4, round(worst_case_distance - overall_distance))

    row += 2
    write(row, 1, "Anteil pref 0")
    write(row, 4, str(round(100 * get_num_assignments_with_preference(0) / len(assignments))) + '%')

    row += 1
    write(row, 1, "Anteil pref 1")
    write(row, 4, str(round(100 * get_num_assignments_with_preference(1) / len(assignments))) + '%')

    row += 1
    write(row, 1, "Anteil pref 2")
    write(row, 4, str(round(100 * get_num_assignments_with_preference(2) / len(assignments))) + '%')

    row += 3
    write(row, 2, "Gewicht")
    write(row, 3, "Kosten")
    write(row, 4, "Ergebnis")

    displayed_weights = [weights[0] + weights[1], weights[2], weights[3]]
    for i, (weight, cost) in enumerate(zip(displayed_weights, get_cost(assignments, worst_case_distance, weights, teachers, desired_workdays, event_durations, event_size))):
        row += 1
        if weight > 0:
            unweighted_cost = str(cost / weight)
        else:
            unweighted_cost = "undefined"
        write(row, 1, "Param. #" + str(i+1))
        write(row, 2, weight)
        write(row, 3, unweighted_cost)
        write(row, 4, cost)

    workbook.save(output_filename)

    print(f"Done. Saved as {output_filename}.")


