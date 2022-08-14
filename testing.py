from ast import AsyncFunctionDef
from openpyxl.styles import Font
import openpyxl
import compute_pulp
import compute_z3
import time
import copy
import json
import random
import compute_pulp2
import statistics

from data import write_model
from data import get_cost


names_file = 'testing/names.json'
input_filename = 'testing/execution times - run 2.xlsx'
solutions_folder = 'testing/10/'
num_iterations = 1
save_assignment_plan = False


# test modes: time, cost or std_deviation
test_mode = "time" 
lambda_step_size = 0.1 # needed if std_deviation is selected


# taken from https://stackoverflow.com/questions/4265546/python-round-to-nearest-05
def round_to(n, precision):
    correction = 0.5 if n >= 0 else -0.5
    return int( n/precision+correction ) * precision


def generate_rand_input(num_teachers, num_events, overlap_chance, workdays_mean, workdays_deviation, event_size_mean, event_size_deviation, event_duration_mean, event_duration_deviation, num_availabilities_mean, num_availabilities_deviation, distance_mean, distance_deviation):
    """ 
    Input: Various problem-size / randomization variables
    Output: The describing variables of a randomized problem
    """
    # teachers
    f = open(names_file)
    json_content = json.load(f)
    first_names = set()
    first_names.update(json_content['girls'] + json_content['boys'])
    all_names = set()

    def char_range(c1, c2):
        """Generates the characters from `c1` to `c2`, inclusive."""
        for c in range(ord(c1), ord(c2)+1):
            yield chr(c)

    for first_name in first_names:
        all_names.update([first_name + ' ' + letter + '.' for letter in char_range('A', 'Z')])

    teachers=[]
    for _ in range(int(num_teachers)):
        teachers.append(all_names.pop())
    f.close()
    
    # events
    events = list(range(int(num_events)))

    # overlapping events
    event_overlap_sets = []
    for event in events:
        num_overlaps = 0
        while(random.random() <= overlap_chance):
            num_overlaps += 1

        if num_overlaps > 0:
            event_overlap_sets.append(set(range(event, min(event + num_overlaps + 1, int(num_events)))))

    # desired workdays
    desired_workdays = {}
    for teacher in teachers:
        min_desired_workdays = 5
        desired_workdays[teacher] = max(min_desired_workdays, round_to(random.gauss(mu=workdays_mean, sigma=workdays_deviation), 0.5))

    # event sizes
    event_size={}
    for event in events:
        event_size[event] = int(max(1, round_to(random.gauss(mu=event_size_mean, sigma=event_size_deviation), 1)))

    # event durations
    event_durations={}
    for event in events:
        event_durations[event] = round_to(random.gauss(mu=event_duration_mean, sigma=event_duration_deviation), 0.5)

    # possible assignments
    possible_assignments = []
    for teacher in teachers:
        availabilities = int(round_to(num_availabilities_mean * (desired_workdays[teacher] / workdays_mean), 1))
        availabilities = min(len(events), max(1, availabilities))
        
        available_events = copy.deepcopy(events)
        selected_events = []
        for _ in range(availabilities):
            selected_events.append(available_events.pop(random.randint(0, len(available_events)-1)))

        for e in selected_events:
            dist = max(0, int(random.gauss(mu=distance_mean, sigma=distance_deviation)))
            prio = random.randint(0,2)
            possible_assignments.append((teacher, e, prio, dist))

    return teachers, events, event_overlap_sets, desired_workdays, event_size, event_durations, possible_assignments



def write_benchmarks(sheet, test_num, column_num, val, color='black'):
    color_codes = {'black':'00000000', 'red':'00FF0000', 'green':'00008000', 'blue':'000000FF', 'orange':'FFA500'}

    # write time
    cell = sheet.cell(row = 6+test_num, column = column_num)
    cell.value = val
    cell.font = Font(color=color_codes[color])



def get_next_input(sheet):
    """
    Returns the data of the first row in input file, which has no execution times yet.
    If an empty row is read, return None's.
    """
    
    # find out the first test case which was not yet executed
    num_rows = 0
    test_num = 0
    for row in sheet.iter_rows(min_row=6):
        if row[0].value is None:
            break
        num_rows += 1 
        if row[21].value is None:
            break
        test_num += 1
            
        
    # return if all tests were executed
    if test_num == num_rows:
        return None, None, None

    # get input data
    randomization_variables = []       # randomization variables
    weights = []
    for i, cell in enumerate(sheet[test_num+6]):
        if i in [3,9,15,20]:
            continue
        if i <= 14:
            randomization_variables.append(float(cell.value))
        elif i <= 19:
            weights.append(float(cell.value))

    return test_num, randomization_variables, weights


def get_std_deviation(assignments, teachers, desired_workdays, event_durations):
    num_assigned_workdays = {t : sum([event_durations[a[1]] for a in assignments if a[0] == t]) for t in teachers}
    absolute_workday_deviations = {t : abs(num_assigned_workdays[t] - desired_workdays[t]) for t in teachers}

    std_deviation = statistics.stdev(absolute_workday_deviations.values())

    return std_deviation



def main():
    workbook = openpyxl.load_workbook(input_filename, data_only=True)
    sheet = workbook.active
    first_row = sheet[4]
    solver_names = list(filter(lambda x: x is not None, list(map(lambda x: x.value, first_row[21:-1]))))

    while(True):
        test_num, rv, weights = get_next_input(sheet)
        if test_num is None:
            break
        
        for k in range(num_iterations):
            # generate random TA problem
            teachers, events, event_overlap_sets, desired_workdays, event_size, event_durations, possible_assignments =\
                generate_rand_input(rv[0], rv[1], rv[2], rv[3], rv[8], rv[4], rv[9], rv[5], rv[10], rv[6], rv[11], rv[7], rv[12])

            print(f"\n\n\n--- TEST #{test_num} ---")
            print('\nsolvers:', solver_names)
            print('\nteachers:', len(teachers))
            print('\nevents:', len(events))
            print('\nevent_overlaps:', event_overlap_sets)
            print('\ndesired_workdays:', desired_workdays)
            print('\nevent sizes:', event_size)
            print('\nevent duration', event_durations)
            print('\noverall available workdays: ',  sum([a * b for a, b in zip(event_size.values(), event_durations.values())]))
            print('\noverall desired workdays: ', sum(desired_workdays.values()))
            sorted_possible_assignments = sorted(possible_assignments, key=lambda a: a[1])
            print('\npossible assignments', sorted_possible_assignments)

            worst_distance = max(1, sum([assignment[3] for assignment in compute_pulp.compute_optimum(teachers, events, event_overlap_sets, desired_workdays, event_size, event_durations, possible_assignments, weights=[0,0,-1,0], worst_case_distance=1)[1]]))

            # compute solution and track time for each solver
            for i, solver_name in enumerate(solver_names):

                if test_mode == "std_deviation":
                    num_rows_to_write = int((1 / lambda_step_size) + 1)
                else:
                    num_rows_to_write = 1

                for row_offset in range(num_rows_to_write):
                    if test_mode == "std_deviation":
                        weights[0] = lambda_step_size * row_offset
                        weights[1] = 1 - (lambda_step_size * row_offset)

                    print(f"\n\n\n--- SOLVING TEST OF ROW {test_num + 6 + row_offset} WITH SOLVER {solver_name} (iteration #{k}) ---\n")
                    start_time = time.time()
                    
                    if solver_name == 'pulp':
                        return_code, assignments = compute_pulp.compute_optimum(teachers, events, event_overlap_sets, desired_workdays, event_size, event_durations, possible_assignments, weights, worst_distance)
                    #elif solver_name == 'pulp2':
                    #   return_code, assignments = compute_pulp2.compute_optimum(teachers, events, event_overlap_sets, desired_workdays, event_size, event_durations, possible_assignments, weights, worst_distance)
                    elif "z3" in solver_name:
                        cost_threshold = float(solver_name.split('-')[1])
                        step_size = float(solver_name.split('-')[2])
                        return_code, assignments = compute_z3.compute_optimum(teachers, events, event_overlap_sets, desired_workdays, event_size, event_durations, possible_assignments, weights, worst_distance,cost_threshold, step_size)
                    else:
                        continue
                
                    if return_code == 1:
                        color = 'black'
                    elif return_code == 0:
                        color = 'orange'
                    elif return_code == -1:
                        color = 'red'
                    else:
                        color = 'blue'

                    column_num = 22 + k + i * (num_iterations + 1)

                    if test_mode == "time":
                        cell_value = time.time() - start_time
                    elif test_mode == "cost":
                        cell_value = sum(get_cost(assignments, worst_distance, weights, teachers, desired_workdays, event_durations, event_size))
                    elif test_mode == "std_deviation":
                        cell_value = get_std_deviation(assignments, teachers, desired_workdays, event_durations)
                    else:
                        print(f"Error, \"{test_mode}\" is not a valid test mode.")
                    
                    if save_assignment_plan:
                        write_model(solutions_folder + "row"+str(test_num+row_offset)+'_'+solver_name+"_"+str(k)+".xlsx", teachers, events, desired_workdays, event_size, event_durations, assignments, weights, worst_distance)
                    
                    write_benchmarks(sheet, test_num+row_offset, column_num, cell_value, color)
        time.sleep(0.5)
        workbook.save(input_filename)
        time.sleep(0.5)

    print("Finished.")



if __name__ == '__main__':
    main()